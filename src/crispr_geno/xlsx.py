"""Write color-coded XLSX versions of the genotype and sequence tables."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


_HF = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HFILL = PatternFill("solid", start_color="305496")
_CEN = Alignment(horizontal="center", vertical="center")
_LFT = Alignment(horizontal="left", vertical="center")
_THIN = Side(border_style="thin", color="000000")
_BD = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_AMB = Side(border_style="medium", color="CC0000")
_BD_AMBIG = Border(left=_AMB, right=_AMB, top=_AMB, bottom=_AMB)
_ARIAL = Font(name="Arial", size=10)
_MONO = Font(name="Arial", size=10)

_GENO_FILLS = {
    "WT":            PatternFill("solid", start_color="D9D9D9"),
    "T":             PatternFill("solid", start_color="C6E0B4"),
    "T/+":           PatternFill("solid", start_color="E2EFDA"),
    "del":           PatternFill("solid", start_color="F8CBAD"),
    "del/+":         PatternFill("solid", start_color="FCE4D6"),
    "T+del":         PatternFill("solid", start_color="FFD966"),
    "T/+ & del/+":   PatternFill("solid", start_color="FFE699"),
    "lowN":          PatternFill("solid", start_color="F4CCCC"),
}


def _style_header(ws) -> None:
    for c in ws[1]:
        c.font = _HF; c.fill = _HFILL; c.alignment = _CEN; c.border = _BD


def genotype_xlsx(header: list[str], rows: list[list[str]], xlsx: str) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "Genotypes"
    ws.append(header)
    for row in rows:
        ws.append(row)
    _style_header(ws)
    for row in ws.iter_rows(min_row=2):
        for i, c in enumerate(row):
            c.font = _ARIAL; c.alignment = _CEN; c.border = _BD
            if i > 0 and c.value in _GENO_FILLS:
                c.fill = _GENO_FILLS[c.value]
    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18
    ws.freeze_panes = "A2"
    wb.save(xlsx)


def _allele_desc_direction(value: str) -> str | None:
    """Classify a per-chromosome allele description (Allele 1 / Allele 2 cell).

    Values look like 'gRNAi:+SEQ', 'gRNAi:-SEQ', or comma-joined when the
    chromosome carries edits at multiple guides ('gRNA1:-AA, gRNA3:+T').
    Returns 'ins' if every edit is an insertion, 'del' if every edit is a
    deletion, None otherwise (mixed, or doesn't match the format).
    """
    parts = [p.strip() for p in value.split(",")]
    directions: set[str] = set()
    for p in parts:
        if not p.startswith("gRNA") or ":" not in p:
            return None
        body = p.split(":", 1)[1]
        if body.startswith("+"):
            directions.add("ins")
        elif body.startswith("-"):
            directions.add("del")
        else:
            return None
    if len(directions) == 1:
        return directions.pop()
    return None


def _pick_seq_fill(value: str) -> PatternFill | None:
    """Pick a fill for a Sequences-sheet cell, or None to leave it unfilled.

    Returns None for free-form text (Notes), numbers (Phased reads), and
    anything that doesn't match a known categorical pattern — so cells
    with non-meaningful content stay uncolored rather than picking up a
    spurious shade that could be misread against the legend.
    """
    if not value:
        return None
    if value == "WT":
        return PatternFill("solid", start_color="D9D9D9")  # mid gray
    if value == "lowN":
        return PatternFill("solid", start_color="F4CCCC")  # light red — no data
    # Plant-level status values
    if value == "Inactive Cas9":
        return PatternFill("solid", start_color="E7E6E6")  # light gray
    if value == "Active Cas9":
        return PatternFill("solid", start_color="E06666")  # saturated red — warning
    # Zygosity calls (blue gradient: lighter = less editing)
    if value == "homozygous":
        return PatternFill("solid", start_color="6FA8DC")  # brighter medium blue
    if value == "biallelic":
        return PatternFill("solid", start_color="EA9999")  # coral
    if value == "heterozygous":
        return PatternFill("solid", start_color="B4C7E7")  # light blue
    if value == "mosaic":
        return PatternFill("solid", start_color="FFE699")  # yellow
    # Structural variant
    if value.startswith("SV:") and value.endswith("/WT"):
        return PatternFill("solid", start_color="D5A6BD")  # light purple (het)
    if value.startswith("SV:") or value == "SV":
        return PatternFill("solid", start_color="B084CB")  # purple (homo SV)
    # Per-guide A1/A2 cells from the phased view: 'X / Y' (always two
    # different slot values — homozygous at a guide collapses to one label).
    if " / " in value:
        a, b = (s.strip() for s in value.split(" / ", 1))
        a_wt = a == "WT"
        b_wt = b == "WT"
        if a_wt or b_wt:
            edited = b if a_wt else a
            if edited == "SV":
                return PatternFill("solid", start_color="D5A6BD")  # SV/WT
            if edited.startswith("+"):
                return PatternFill("solid", start_color="E2EFDA")  # het, ins
            if edited.startswith("-"):
                return PatternFill("solid", start_color="FCE4D6")  # het, del
            return None
        # both edited, different lesions → biallelic at this guide
        if a == "SV" or b == "SV":
            return PatternFill("solid", start_color="B084CB")
        return PatternFill("solid", start_color="EA9999")  # biallelic coral
    # Legacy per-guide notation: 'X/WT' (het) and 'A/B' (biallelic, no spaces)
    if value.endswith("/WT"):
        if value.startswith("+"):
            return PatternFill("solid", start_color="E2EFDA")  # het, ins
        if value.startswith("-"):
            return PatternFill("solid", start_color="FCE4D6")  # het, del
        return None
    if "/" in value:
        return PatternFill("solid", start_color="EA9999")  # biallelic coral
    # Per-guide homozygous slot labels
    if value.startswith("+"):
        return PatternFill("solid", start_color="C6E0B4")  # green, ins
    if value.startswith("-"):
        return PatternFill("solid", start_color="F8CBAD")  # peach, del
    # Allele 1 / Allele 2 per-chromosome descriptions ('gRNAi:+SEQ' etc.)
    direction = _allele_desc_direction(value)
    if direction == "ins":
        return PatternFill("solid", start_color="C6E0B4")
    if direction == "del":
        return PatternFill("solid", start_color="F8CBAD")
    return None


_LEGEND_LEGACY = [
    ("Inactive Cas9", "Status: all sites resolve cleanly — no current editing signature in this amplicon",
        PatternFill("solid", start_color="E7E6E6")),
    ("Active Cas9",   "Status: at least one site shows mosaic / residual editing — ongoing Cas9 activity",
        PatternFill("solid", start_color="E06666")),
    ("WT",       "wildtype — no edit detected at this site",
        PatternFill("solid", start_color="D9D9D9")),
    ("lowN",     "fewer than 10 reads spanning the cut — cannot call",
        PatternFill("solid", start_color="F4CCCC")),
    ("+XYZ",     "homozygous insertion of XYZ at cut site",
        PatternFill("solid", start_color="C6E0B4")),
    ("+XYZ/WT",  "heterozygous: one chromosome has +XYZ, the other is wildtype",
        PatternFill("solid", start_color="E2EFDA")),
    ("-XYZ",     "homozygous deletion of XYZ from reference",
        PatternFill("solid", start_color="F8CBAD")),
    ("-XYZ/WT",  "heterozygous: one chromosome has -XYZ, the other is wildtype",
        PatternFill("solid", start_color="FCE4D6")),
    ("A/B",      "biallelic: one chromosome has allele A, the other has a different mutant allele B",
        PatternFill("solid", start_color="EA9999")),
    ("SV:-Nbp",  "homozygous large deletion between two cut sites (N bp removed)",
        PatternFill("solid", start_color="B084CB")),
    ("SV:-Nbp/WT", "heterozygous large deletion between two cut sites",
        PatternFill("solid", start_color="D5A6BD")),
    ("ambiguous (cell border)", "competing allele present at ≥15% — check editing_rates.csv (shown as a red border on the cell)",
        None),
]


_LEGEND_PHASED = [
    # Zygosity calls (Zygosity call column)
    ("WT",           "no edits — both chromosomes match the reference",
        PatternFill("solid", start_color="D9D9D9")),
    ("heterozygous", "one chromosome edited, the other wildtype",
        PatternFill("solid", start_color="B4C7E7")),
    ("homozygous",   "same edit on both chromosomes — fixed, breeds true",
        PatternFill("solid", start_color="6FA8DC")),
    ("biallelic",    "different edit on each chromosome — segregates in progeny",
        PatternFill("solid", start_color="EA9999")),
    ("mosaic",       "3+ distinct haplotypes — Cas9 likely still active, or chimeric tissue",
        PatternFill("solid", start_color="FFE699")),
    ("lowN",         "too few reads spanning the cut(s) to make a call",
        PatternFill("solid", start_color="F4CCCC")),
    # Plant-level status (Status column)
    ("Inactive Cas9", "Status column: all guides resolve cleanly — no current editing signature",
        PatternFill("solid", start_color="E7E6E6")),
    ("Active Cas9",   "Status column: at least one guide shows ongoing editing",
        PatternFill("solid", start_color="E06666")),
    # Per-chromosome allele notation (Allele 1 / Allele 2 columns)
    ("SV:-Nbp",      "Allele 1/2: large deletion of N bp spanning two guide cuts",
        PatternFill("solid", start_color="B084CB")),
    ("gRNAi:+SEQ",   "Allele 1/2: insertion of SEQ at guide i's cut site",
        PatternFill("solid", start_color="C6E0B4")),
    ("gRNAi:-SEQ",   "Allele 1/2: deletion of SEQ at guide i's cut site",
        PatternFill("solid", start_color="F8CBAD")),
    # Cell indicator
    ("ambiguous (cell border)", "competing minor allele present — see haplotypes.csv (cell shown with a red border)",
        None),
]


def _write_sheet(ws, header, rows, conf_rows, legend_entries):
    ws.append(header)
    for row in rows:
        ws.append(row)
    _style_header(ws)

    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        row[0].font = _ARIAL; row[0].alignment = _CEN; row[0].border = _BD
        conf = conf_rows[ridx] if conf_rows and ridx < len(conf_rows) else None
        for ci, c in enumerate(row[1:]):
            c.font = _MONO; c.alignment = _LFT
            c.border = _BD_AMBIG if (conf and ci < len(conf) and conf[ci] == "ambiguous") else _BD
            # Allele 1 / Allele 2 are free-text per-chromosome descriptions —
            # leave uncolored. Coloring there was misleading (descriptive
            # text shouldn't read as a categorical fill).
            header_val = ws.cell(1, ci + 2).value
            if header_val in ("Allele 1", "Allele 2"):
                continue
            fill = _pick_seq_fill(c.value or "")
            if fill is not None:
                c.fill = fill

    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, ws.max_column + 1):
        header_val = ws.cell(1, col_idx).value
        if header_val == "Notes":
            width = 42
        elif header_val == "Phased reads":
            width = 14
        elif header_val in ("Status", "Zygosity call"):
            width = 12
        else:
            width = 28
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    last_col = ws.max_column
    buffer_col = last_col + 1
    color_col = last_col + 2
    label_col = last_col + 3
    meaning_col = last_col + 4

    for col, title in ((color_col, "Color"),
                       (label_col, "Allele / Notation"),
                       (meaning_col, "Meaning")):
        cell = ws.cell(1, col, value=title)
        cell.font = _HF; cell.fill = _HFILL; cell.alignment = _CEN; cell.border = _BD

    row_idx = 2
    for label, meaning, fill in legend_entries:
        if fill is not None:
            ws.cell(row_idx, color_col).fill = fill
            ws.cell(row_idx, color_col).border = _BD
        elif label.endswith("(cell border)"):
            ws.cell(row_idx, color_col).border = _BD_AMBIG
        lbl = ws.cell(row_idx, label_col, value=label)
        lbl.font = _ARIAL; lbl.alignment = _LFT
        mng = ws.cell(row_idx, meaning_col, value=meaning)
        mng.font = _ARIAL; mng.alignment = _LFT
        for col in (label_col, meaning_col):
            ws.cell(row_idx, col).border = _BD
        row_idx += 1

    ws.column_dimensions[ws.cell(1, buffer_col).column_letter].width = 3
    ws.column_dimensions[ws.cell(1, color_col).column_letter].width = 8
    ws.column_dimensions[ws.cell(1, label_col).column_letter].width = 20
    ws.column_dimensions[ws.cell(1, meaning_col).column_letter].width = 70


def sequence_xlsx(
    header: list[str],
    rows: list[list[str]],
    xlsx: str,
    conf_rows: list[list[str]] | None = None,
    secondary_header: list[str] | None = None,
    secondary_rows: list[list[str]] | None = None,
    secondary_conf_rows: list[list[str]] | None = None,
    secondary_title: str = "Per-guide raw",
    primary_is_phased: bool = False,
) -> None:
    """Write the results workbook.

    The primary sheet ('Sequences') receives `header` / `rows`. If a secondary
    sheet is provided, it is added as a hidden sheet for downstream auditing.
    `primary_is_phased` selects which legend palette goes on the primary sheet.
    """
    wb = Workbook(); ws = wb.active; ws.title = "Sequences"
    primary_legend = _LEGEND_PHASED if primary_is_phased else _LEGEND_LEGACY
    _write_sheet(ws, header, rows, conf_rows, primary_legend)

    if secondary_header is not None and secondary_rows is not None:
        ws2 = wb.create_sheet(title=secondary_title)
        secondary_legend = _LEGEND_LEGACY if primary_is_phased else _LEGEND_PHASED
        _write_sheet(ws2, secondary_header, secondary_rows,
                     secondary_conf_rows, secondary_legend)
        ws2.sheet_state = "hidden"

    wb.save(xlsx)

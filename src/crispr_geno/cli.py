"""Command-line interface for crispr-geno."""
from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from typing import Optional

from . import __version__
from .analysis import (
    AlleleCall,
    Guide,
    GuideCounts,
    align,
    call_allele,
    check_tools,
    count_indels,
    find_large_deletions,
    resolve_guides,
)
from .xlsx import sequence_xlsx


@dataclass
class Sample:
    id: str
    fastq: str


# ---------------------------------------------------------------- interactive

def _read_multiline(prompt: str) -> str:
    """Read a multiline sequence from stdin. Finish by entering an empty line
    or EOF (Ctrl-D)."""
    print(prompt)
    print("  (paste, then press Enter on an empty line to finish)")
    lines = []
    while True:
        try:
            line = input()
        except EOFError:
            break
        if not line.strip():
            break
        lines.append(line.strip())
    return "".join(lines).upper()


def prompt_guides() -> list[Guide]:
    print()
    print("Enter guide sequences one per line as 'NAME SEQUENCE' (whitespace-separated).")
    print("Example:")
    print("  gRNA1 AAGCTATTCTCCCTCGGGGT")
    print("  gRNA2 TCACACGAGCGTAACAAGGA")
    print("(press Enter on an empty line to finish)")
    guides = []
    i = 1
    while True:
        try:
            line = input().strip()
        except EOFError:
            break
        if not line:
            break
        parts = line.split()
        if len(parts) == 1:
            # guide sequence only — auto-name
            name, seq = f"gRNA{i}", parts[0]
        elif len(parts) >= 2:
            name, seq = parts[0], parts[1]
        else:
            print(f"  skipping malformed line: {line!r}")
            continue
        guides.append(Guide(name=name, sequence=seq.upper()))
        i += 1
    if not guides:
        sys.exit("no guides given")
    return guides


def prompt_samples_dir() -> tuple[str, list[Sample]]:
    print()
    directory = input("Path to directory containing fastq files: ").strip().strip("'\"")
    directory = os.path.expanduser(directory)
    if not os.path.isdir(directory):
        sys.exit(f"not a directory: {directory}")
    fastqs = sorted(f for f in os.listdir(directory) if f.endswith((".fastq", ".fq", ".fastq.gz", ".fq.gz")))
    if not fastqs:
        sys.exit(f"no fastq files found in {directory}")
    samples = [Sample(id=os.path.splitext(os.path.splitext(f)[0])[0] if f.endswith(".gz")
                      else os.path.splitext(f)[0], fastq=os.path.join(directory, f))
               for f in fastqs]
    print(f"  found {len(samples)} fastq files")
    return directory, samples


# ----------------------------------------------------------- CLI-arg parsing

def parse_guide_arg(arg: str, idx: int) -> Guide:
    """--guide NAME:SEQ or --guide SEQ (auto-names gRNA{idx})."""
    if ":" in arg:
        name, seq = arg.split(":", 1)
    else:
        name, seq = f"gRNA{idx}", arg
    return Guide(name=name.strip(), sequence=seq.strip().upper())


def load_samples_from_dir(directory: str) -> list[Sample]:
    fastqs = sorted(f for f in os.listdir(directory)
                    if f.endswith((".fastq", ".fq", ".fastq.gz", ".fq.gz")))
    out = []
    for f in fastqs:
        stem = f
        for ext in (".fastq.gz", ".fq.gz", ".fastq", ".fq"):
            if stem.endswith(ext):
                stem = stem[:-len(ext)]
                break
        out.append(Sample(id=stem, fastq=os.path.join(directory, f)))
    return out


def load_samples_tsv(path: str) -> list[Sample]:
    """Simple TSV: `sample_id<TAB>fastq_path`, with or without a header row.
    Paths are resolved relative to the TSV's directory."""
    out = []
    base = os.path.dirname(os.path.abspath(path))
    with open(path) as f:
        for i, line in enumerate(f):
            parts = line.rstrip("\n").replace("\r", "").split("\t")
            if len(parts) < 2:
                continue
            if i == 0 and parts[0].lower() in ("sample", "sample_id", "name", "id"):
                continue
            fq = parts[1] if os.path.isabs(parts[1]) else os.path.join(base, parts[1])
            out.append(Sample(id=parts[0], fastq=fq))
    return out


def _iter_samples_rows(path: str):
    """Yield list-of-strings rows from a samples sheet. Auto-detects format
    by extension: .xlsx -> openpyxl, .csv -> comma-split, else TSV."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield ["" if v is None else str(v) for v in row]
    else:
        delim = "," if ext == ".csv" else "\t"
        with open(path) as f:
            for raw in f:
                yield raw.rstrip("\n").replace("\r", "").split(delim)


def load_samples_crispresso(path: str) -> tuple[str, list[Guide], list[Sample]]:
    """Load a CRISPResso-style batch sheet (TSV/CSV/XLSX) with columns:
        name, fastq_r1, amplicon_seq, guide_seq
    where guide_seq is comma-separated guide sequences.

    Returns (amplicon, guides, samples). Amplicon + guides are taken from the
    first data row (all rows are assumed to share the same amplicon/guides).
    Fastq paths are resolved relative to the samples sheet's directory.
    """
    base = os.path.dirname(os.path.abspath(path))
    samples: list[Sample] = []
    amplicon: str | None = None
    guides: list[Guide] = []
    header_map: dict[str, int] = {}

    for i, parts in enumerate(_iter_samples_rows(path)):
        if not parts or not parts[0].strip():
            continue
        if i == 0:
            header_map = {h.strip().lower(): idx for idx, h in enumerate(parts)}
            required = {"name", "fastq_r1", "amplicon_seq", "guide_seq"}
            if not required.issubset(header_map):
                raise ValueError(
                    f"{path}: expected CRISPResso columns "
                    f"(name, fastq_r1, amplicon_seq, guide_seq); "
                    f"got {parts}"
                )
            continue

        name = parts[header_map["name"]].strip()
        fq_rel = parts[header_map["fastq_r1"]].strip().strip('"\'')
        amp_row = parts[header_map["amplicon_seq"]].strip().upper()
        guide_row = parts[header_map["guide_seq"]].strip().strip('"\'')

        if not name or not fq_rel:
            continue

        fq = fq_rel if os.path.isabs(fq_rel) else os.path.join(base, fq_rel)
        samples.append(Sample(id=name, fastq=fq))

        if amplicon is None and amp_row:
            amplicon = amp_row
        if not guides and guide_row:
            for j, g in enumerate(
                s.strip().upper() for s in guide_row.split(",") if s.strip()
            ):
                guides.append(Guide(name=f"gRNA{j + 1}", sequence=g))

    if amplicon is None:
        raise ValueError(f"{path}: no amplicon sequence found in any row")
    if not guides:
        raise ValueError(f"{path}: no guides found in any row")
    return amplicon, guides, samples


def _has_crispresso_schema(path: str) -> bool:
    """Peek at the first row of a file and check if it matches the CRISPResso
    batch schema (name, fastq_r1, amplicon_seq, guide_seq in the header)."""
    required = {"name", "fastq_r1", "amplicon_seq", "guide_seq"}
    try:
        for row in _iter_samples_rows(path):
            if not row:
                continue
            header = {str(h).strip().lower() for h in row}
            return required.issubset(header)
    except Exception:
        return False
    return False


def find_crispresso_samples_file(directory: str, parents_to_check: int = 2) -> str | None:
    """Find a CRISPResso-style samples sheet. Looks in the given directory and
    up to `parents_to_check` parent directories. Preference order within each
    directory:
       1. canonical filenames (samples.txt/.tsv/.csv/.xlsx)
       2. any .xlsx / .csv / .tsv / .txt whose header row matches the
          CRISPResso schema (name, fastq_r1, amplicon_seq, guide_seq).
    """
    canonical = ("samples.txt", "samples.tsv", "samples.csv", "samples.xlsx")
    search_exts = (".xlsx", ".csv", ".tsv", ".txt")

    d = os.path.abspath(directory)
    for _ in range(parents_to_check + 1):
        # 1. canonical names first
        for name in canonical:
            p = os.path.join(d, name)
            if os.path.isfile(p):
                return p
        # 2. any file with the CRISPResso schema
        try:
            for name in sorted(os.listdir(d)):
                if name.startswith("."):
                    continue
                if not name.lower().endswith(search_exts):
                    continue
                p = os.path.join(d, name)
                if os.path.isfile(p) and _has_crispresso_schema(p):
                    return p
        except OSError:
            pass
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    return None


# ------------------------------------------------------------------- driver

def run_analysis(amplicon: str, guides: list[Guide], samples: list[Sample],
                 out_dir: str, window: int, min_spanning: int,
                 homo_threshold: float, het_threshold: float,
                 min_read_len: int, threads: int) -> None:
    check_tools()
    os.makedirs(out_dir, exist_ok=True)
    bam_dir = os.path.join(out_dir, "bam")
    os.makedirs(bam_dir, exist_ok=True)

    # write reference.fasta
    ref_fa = os.path.join(out_dir, "reference.fasta")
    with open(ref_fa, "w") as f:
        f.write(">amplicon\n")
        f.write(amplicon.upper() + "\n")
    ref_name = "amplicon"
    ref_seq = amplicon.upper()

    # resolve guides
    guides = resolve_guides(ref_seq, guides)
    if not guides:
        sys.exit("no guides match the amplicon — check orientation and sequences")
    print("Resolved guide cut positions:")
    for g in guides:
        print(f"  {g.name}  {g.sequence}  strand={g.strand}  cut@{g.cut_pos}")

    # Accumulate results in memory so we can write CSV + XLSX from the same source.
    rates_header = [
        "sample", "guide", "guide_seq", "cut_pos", "spanning_reads",
        "ins1_pct", "del1_pct", "ins_ge2_pct", "del_ge2_pct",
        "any_indel_pct", "ge2_indel_pct",
        "ins1_n", "del1_n", "ins_ge2_n", "del_ge2_n",
        "call", "top_allele_frac", "second_allele_frac", "wt_frac",
        "call_confidence", "plant_status",
    ]
    rates_rows: list[list[str]] = []
    seq_rows: list[list[str]] = []
    # Parallel matrix of confidence strings, one per cell of seq_rows.
    conf_rows: list[list[str]] = []
    # Cache SV detection results per plant so we can emit rows/columns later.
    sv_results_per_plant: list[tuple[str, dict]] = []  # (plant_id, results_dict)
    sv_rates_rows: list[list[str]] = []

    for s in samples:
        if not os.path.exists(s.fastq):
            print(f"  SKIP {s.id}: {s.fastq} not found")
            continue
        bam = os.path.join(bam_dir, f"{s.id}.bam")
        if not os.path.exists(bam) or not os.path.exists(bam + ".bai"):
            print(f"aligning {s.id}...")
            align(ref_fa, s.fastq, bam, threads=threads)

        # Collect calls for this plant so we can assign a per-plant status after.
        plant_calls: list[AlleleCall] = []
        plant_guide_rates_rows: list[list[str]] = []
        for g in guides:
            c = count_indels(bam, ref_name, g.cut_pos, window, min_read_len)
            ac = call_allele(bam, ref_name, ref_seq, g.cut_pos, window,
                             min_spanning, homo_threshold, het_threshold,
                             min_read_len)
            any_n = min(c.spanning, c.ins1 + c.del1 + c.ins_ge2 + c.del_ge2)
            ge2_n = min(c.spanning, c.ins_ge2 + c.del_ge2)
            plant_guide_rates_rows.append([
                s.id, g.name, g.sequence, str(g.cut_pos), str(c.spanning),
                f"{c.pct('ins1'):.2f}", f"{c.pct('del1'):.2f}",
                f"{c.pct('ins_ge2'):.2f}", f"{c.pct('del_ge2'):.2f}",
                f"{100.0 * any_n / c.spanning if c.spanning else 0:.2f}",
                f"{100.0 * ge2_n / c.spanning if c.spanning else 0:.2f}",
                str(c.ins1), str(c.del1), str(c.ins_ge2), str(c.del_ge2),
                ac.call,
                f"{ac.top_frac:.2f}", f"{ac.second_frac:.2f}", f"{ac.wt_frac:.2f}",
                ac.confidence,
            ])
            plant_calls.append(ac)
            print(f"  {s.id} {g.name}: span={c.spanning} "
                  f"ins1={c.ins1} del1={c.del1} ins_ge2={c.ins_ge2} del_ge2={c.del_ge2} "
                  f"-> {ac.call} [{ac.confidence}]")

        # Structural-variant detection: large deletions spanning two cut sites.
        sv_results = find_large_deletions(
            bam, ref_name, guides,
            min_size=50, breakpoint_window=20,
            min_read_len=min_read_len,
            homo_threshold=homo_threshold,
            het_threshold=het_threshold,
            min_supporting=5,
        )
        sv_results_per_plant.append((s.id, sv_results))
        for pair, info in sv_results.items():
            i, j = pair
            sv_rates_rows.append([
                s.id,
                f"{guides[i].name}-{guides[j].name}",
                f"{guides[i].cut_pos}-{guides[j].cut_pos}",
                str(info["spanning_reads"]),
                str(info["supporting_reads"]),
                str(info["median_size"]),
                f"{info['top_frac']:.2f}",
                info["call"],
            ])

        # Per-plant status:
        #   "Cas9+"  : at least one site shows the mosaic / residual-editing
        #              signature (any ambiguous call)
        #   "stable" : all sites are clean (WT / homo / het / lowN)
        plant_status = "Cas9+" if any(ac.confidence == "ambiguous"
                                       for ac in plant_calls) else "stable"

        # Append Status column to every rates row for this plant
        for r in plant_guide_rates_rows:
            r.append(plant_status)
            rates_rows.append(r)

        # Per-plant base row (Status + per-guide calls). SV columns are appended
        # below, after we know which pairs to include.
        seq_rows.append([s.id, plant_status] + [ac.call for ac in plant_calls])
        conf_rows.append(["ambiguous" if plant_status == "Cas9+" else ""] +
                         [ac.confidence for ac in plant_calls])

    # Figure out which (i,j) guide pairs had any SV detected across the dataset.
    # Only those pairs get columns. Pairs with zero detections in any plant are
    # omitted to keep the table narrow.
    sv_pairs_seen: list[tuple[int, int]] = []
    for _, res in sv_results_per_plant:
        for pair, info in res.items():
            # include pair only if the call is non-trivial (not 'WT' and not 'lowN')
            if info["call"] not in ("WT", "lowN") and pair not in sv_pairs_seen:
                sv_pairs_seen.append(pair)
    sv_pairs_seen.sort()

    # Build the full header now that we know SV pairs.
    table_header = ["Plant ID", "Status"] + [f"gRNA{i+1}" for i in range(len(guides))]
    for (i, j) in sv_pairs_seen:
        table_header.append(f"{guides[i].name}↔{guides[j].name} SV")

    # Append SV cells to each plant's row in the same order.
    for row_idx, (sid, res) in enumerate(sv_results_per_plant):
        for pair in sv_pairs_seen:
            info = res.get(pair)
            if info is None:
                seq_rows[row_idx].append("WT")
                conf_rows[row_idx].append("")
            else:
                seq_rows[row_idx].append(info["call"])
                # SV cells don't get an ambiguous border from this pipeline yet
                conf_rows[row_idx].append("")

    # Write outputs: rates as CSV; sequence table as XLSX (with Legend sheet).
    import csv as _csv

    def _csv_safe(v: str) -> str:
        # Excel interprets cells starting with =, +, -, or @ as formulas, which
        # turns '+T/WT' into a #NAME? error. Prefix those with a tab so Excel
        # treats them as plain text; Excel strips the tab on import.
        if isinstance(v, str) and v[:1] in "=+-@":
            return "\t" + v
        return v

    rates_csv = os.path.join(out_dir, "editing_rates.csv")
    with open(rates_csv, "w", newline="") as f:
        w = _csv.writer(f)
        w.writerow(rates_header)
        for row in rates_rows:
            w.writerow([_csv_safe(c) for c in row])

    # Separate SV events CSV (one row per detected large deletion per plant)
    if sv_rates_rows:
        sv_csv = os.path.join(out_dir, "sv_events.csv")
        with open(sv_csv, "w", newline="") as f:
            w = _csv.writer(f)
            w.writerow([
                "sample", "guides", "cut_positions", "spanning_reads",
                "supporting_reads", "median_deletion_size_bp",
                "supporting_frac", "call",
            ])
            for row in sv_rates_rows:
                w.writerow([_csv_safe(c) for c in row])

    print("\nWriting xlsx...")
    seq_xlsx_path = os.path.join(out_dir, "results.xlsx")
    sequence_xlsx(table_header, seq_rows, seq_xlsx_path, conf_rows=conf_rows)

    print(f"\nDone.  Outputs in {out_dir}/:")
    for name in ("editing_rates.csv", "results.xlsx"):
        print(f"  {name}")
    if sv_rates_rows:
        print("  sv_events.csv")


# ---------------------------------------------------------------- main CLI

def main(argv: Optional[list[str]] = None) -> None:
    p = argparse.ArgumentParser(
        prog="crispr-geno",
        description="Genotype long-read CRISPR amplicon sequencing.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--version", action="version", version=f"crispr-geno {__version__}")

    # Amplicon + guides: either provide on cli, or get prompted interactively
    p.add_argument("--amplicon", help="Amplicon sequence (or path to a FASTA file).")
    p.add_argument("--guide", action="append", default=[],
                   metavar="NAME:SEQ",
                   help="Guide (repeatable). 'NAME:SEQ' or just 'SEQ'. "
                        "e.g. --guide gRNA1:AAGCTATTCTCCCTCGGGGT")

    # Samples: directory of fastqs, OR a tsv with id<TAB>path
    p.add_argument("--input-dir", help="Directory containing .fastq files (one per sample).")
    p.add_argument("--samples-tsv", help="TSV with 'sample_id<TAB>fastq_path' per row.")

    p.add_argument("-o", "--output", default="crispr_geno_output",
                   help="Output directory.")

    # Thresholds / parameters
    p.add_argument("--window", type=int, default=5,
                   help="bp window on each side of the cut site.")
    p.add_argument("--min-spanning", type=int, default=20,
                   help="Minimum spanning reads required to call a genotype.")
    p.add_argument("--homo-threshold", type=float, default=70.0,
                   help="Percent threshold for a homozygous call.")
    p.add_argument("--het-threshold", type=float, default=20.0,
                   help="Percent threshold for a heterozygous call.")
    p.add_argument("--min-read-len", type=int, default=200,
                   help="Minimum read length (bp) to include.")
    p.add_argument("--threads", type=int, default=4,
                   help="Threads for minimap2 / samtools.")

    args = p.parse_args(argv)

    # ---- Try auto-detection of a CRISPResso-style samples.txt first.
    # If the user only points at a directory and doesn't pass amplicon/guides,
    # a samples.txt in that directory provides everything we need.
    amplicon: str | None = None
    guides: list[Guide] | None = None
    samples: list[Sample] | None = None

    auto_path = args.samples_tsv
    if auto_path is None:
        # look in the input dir they pointed at (and its parents), else prompt.
        if args.input_dir and os.path.isdir(args.input_dir):
            auto_path = find_crispresso_samples_file(args.input_dir)
            if auto_path is None:
                print(f"\nNo samples sheet (samples.txt/.tsv/.csv/.xlsx) found in:")
                print(f"  {args.input_dir}")
                print(f"  or up to 2 parent directories")
                print(f"You'll be prompted for the amplicon and guides instead.")
        if auto_path is None and args.amplicon is None and not args.guide and not args.input_dir:
            # nothing at all given — ask where the fastqs are and look there
            input_dir, _dir_samples = prompt_samples_dir()
            args.input_dir = input_dir
            auto_path = find_crispresso_samples_file(input_dir)

    if auto_path and not args.amplicon and not args.guide:
        # try CRISPResso format
        try:
            amp, gds, smps = load_samples_crispresso(auto_path)
            amplicon = amp
            guides = gds
            samples = smps
            print(f"\nLoaded {auto_path}")
            print(f"  {len(samples)} samples")
            print(f"  amplicon: {len(amplicon)} bp")
            print(f"  guides:   {', '.join(g.sequence for g in guides)}")
        except ValueError as e:
            print(f"  (not a CRISPResso batch file: {e})")

    # ---- resolve amplicon (if not already)
    if amplicon is None:
        if args.amplicon is None:
            amplicon = _read_multiline(
                "\nPaste the amplicon sequence (ACGT characters only — line breaks OK):"
            )
            if not amplicon:
                sys.exit("no amplicon given")
        elif os.path.isfile(args.amplicon):
            with open(args.amplicon) as f:
                f.readline() if f.read(1) == ">" else f.seek(0)
                amplicon = "".join(
                    line.strip() for line in f if not line.startswith(">")
                ).upper()
        else:
            amplicon = args.amplicon.upper().replace(" ", "").replace("\n", "")

    # ---- resolve guides (if not already)
    if guides is None:
        if args.guide:
            guides = [parse_guide_arg(g, i + 1) for i, g in enumerate(args.guide)]
        else:
            guides = prompt_guides()

    # ---- resolve samples (if not already)
    if samples is None:
        if args.samples_tsv:
            samples = load_samples_tsv(args.samples_tsv)
        elif args.input_dir:
            samples = load_samples_from_dir(args.input_dir)
        else:
            _, samples = prompt_samples_dir()
    if not samples:
        sys.exit("no samples found")

    run_analysis(
        amplicon=amplicon,
        guides=guides,
        samples=samples,
        out_dir=args.output,
        window=args.window,
        min_spanning=args.min_spanning,
        homo_threshold=args.homo_threshold,
        het_threshold=args.het_threshold,
        min_read_len=args.min_read_len,
        threads=args.threads,
    )


if __name__ == "__main__":
    main()
